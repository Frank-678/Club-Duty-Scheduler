import argparse
import json
import os
import uuid
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from flask import Flask, jsonify, render_template, request, send_from_directory
except ModuleNotFoundError:
    Flask = None
    jsonify = None
    render_template = None
    request = None
    send_from_directory = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    Workbook = None
    Alignment = None
    Font = None
    PatternFill = None
    get_column_letter = None

DAYS = ["1", "2", "3", "4", "5", "6", "7"]
SLOTS = ["1", "2", "3", "4", "5", "6"]
DAY_NAME = {"1": "周一", "2": "周二", "3": "周三", "4": "周四", "5": "周五", "6": "周六", "7": "周日"}
PRIORITY_COST = {"prefer_more": -20, "normal": 0, "prefer_less": 20}
REPEAT_LAYER_COST = 1000


class ValidationError(Exception):
    pass


@dataclass
class Edge:
    to: int
    rev: int
    cap: int
    cost: int


class MinCostMaxFlow:
    def __init__(self, n: int) -> None:
        self.n = n
        self.graph: List[List[Edge]] = [[] for _ in range(n)]

    def add_edge(self, fr: int, to: int, cap: int, cost: int) -> None:
        fwd = Edge(to=to, rev=len(self.graph[to]), cap=cap, cost=cost)
        rev = Edge(to=fr, rev=len(self.graph[fr]), cap=0, cost=-cost)
        self.graph[fr].append(fwd)
        self.graph[to].append(rev)

    def min_cost_max_flow(self, s: int, t: int) -> Tuple[int, int]:
        flow = 0
        cost = 0
        while True:
            inf = 10**18
            dist = [inf] * self.n
            in_queue = [False] * self.n
            prev_v = [-1] * self.n
            prev_e = [-1] * self.n
            dist[s] = 0
            queue = [s]
            in_queue[s] = True
            head = 0
            while head < len(queue):
                v = queue[head]
                head += 1
                in_queue[v] = False
                for i, e in enumerate(self.graph[v]):
                    if e.cap <= 0:
                        continue
                    nd = dist[v] + e.cost
                    if nd < dist[e.to]:
                        dist[e.to] = nd
                        prev_v[e.to] = v
                        prev_e[e.to] = i
                        if not in_queue[e.to]:
                            queue.append(e.to)
                            in_queue[e.to] = True
            if dist[t] == inf:
                break
            addf = inf
            v = t
            while v != s:
                u = prev_v[v]
                e = self.graph[u][prev_e[v]]
                addf = min(addf, e.cap)
                v = u
            v = t
            while v != s:
                u = prev_v[v]
                ei = prev_e[v]
                e = self.graph[u][ei]
                e.cap -= addf
                self.graph[v][e.rev].cap += addf
                v = u
            flow += addf
            cost += addf * dist[t]
        return flow, cost


def empty_day(value: int = 0) -> Dict[str, int]:
    return {slot: value for slot in SLOTS}


def default_open_shifts() -> Dict[str, Dict[str, int]]:
    return {day: empty_day(1) for day in DAYS}


def normalize_int(value, default: int, field_name: str, min_value: int = 1, max_value: Optional[int] = None) -> int:
    if value is None:
        value = default
    try:
        result = int(value)
    except (TypeError, ValueError):
        raise ValidationError(f"{field_name} 必须是整数")
    if result < min_value:
        raise ValidationError(f"{field_name} 必须 >= {min_value}")
    if max_value is not None and result > max_value:
        raise ValidationError(f"{field_name} 必须 <= {max_value}")
    return result


def normalize_input(data: dict) -> dict:
    if not isinstance(data, dict):
        raise ValidationError("输入必须是 JSON 对象")

    config = deepcopy(data.get("config", {}))
    open_shifts = config.get("open_shifts", default_open_shifts())
    max_shifts_per_member = normalize_int(
        config.get("max_shifts_per_member", 1),
        default=1,
        field_name="config.max_shifts_per_member",
        min_value=1,
        max_value=len(DAYS) * len(SLOTS),
    )
    members = data.get("members")

    if not isinstance(members, list) or not members:
        raise ValidationError("members 必须是非空数组")

    if not isinstance(open_shifts, dict):
        raise ValidationError("config.open_shifts 必须是对象")

    normalized_open_shifts: Dict[str, Dict[str, int]] = {}
    for day in DAYS:
        if day not in open_shifts or not isinstance(open_shifts[day], dict):
            raise ValidationError(f"config.open_shifts 缺少 '{day}'")
        normalized_open_shifts[day] = {}
        for slot in SLOTS:
            if slot not in open_shifts[day]:
                raise ValidationError(f"config.open_shifts['{day}'] 缺少 '{slot}'")
            value = open_shifts[day][slot]
            if value not in (0, 1):
                raise ValidationError(f"config.open_shifts 在 day={day}, slot={slot} 的值必须是 0 或 1")
            normalized_open_shifts[day][slot] = int(value)

    seen_names = set()
    normalized_members = []
    for member in members:
        if not isinstance(member, dict):
            raise ValidationError("每个成员都必须是对象")
        name = str(member.get("name", "")).strip()
        if not name:
            raise ValidationError("每个成员都必须提供非空 name")
        if name in seen_names:
            raise ValidationError(f"成员姓名重复：{name}")
        seen_names.add(name)

        priority = member.get("priority", "normal")
        if priority not in PRIORITY_COST:
            raise ValidationError(f"成员 {name} 的 priority 必须是 prefer_more / normal / prefer_less")

        ban_days = member.get("ban_days", [])
        if not isinstance(ban_days, list):
            raise ValidationError(f"成员 {name} 的 ban_days 必须是数组")
        ban_days = [str(day) for day in ban_days]
        for day in ban_days:
            if day not in DAYS:
                raise ValidationError(f"成员 {name} 的 ban_days 包含非法日期：{day}")

        schedule = member.get("schedule")
        if not isinstance(schedule, dict):
            raise ValidationError(f"成员 {name} 的 schedule 必须是对象")
        normalized_schedule: Dict[str, Dict[str, int]] = {}
        for day in DAYS:
            if day not in schedule or not isinstance(schedule[day], dict):
                raise ValidationError(f"成员 {name} 缺少 schedule['{day}']")
            normalized_schedule[day] = {}
            for slot in SLOTS:
                if slot not in schedule[day]:
                    raise ValidationError(f"成员 {name} 缺少 schedule['{day}']['{slot}']")
                value = schedule[day][slot]
                if value not in (0, 1):
                    raise ValidationError(f"成员 {name} 在 day={day}, slot={slot} 的值必须是 0 或 1")
                normalized_schedule[day][slot] = int(value)

        normalized_members.append(
            {
                "name": name,
                "priority": priority,
                "ban_days": ban_days,
                "schedule": normalized_schedule,
            }
        )

    normalized = {
        "config": {
            "open_shifts": normalized_open_shifts,
            "max_shifts_per_member": max_shifts_per_member,
        },
        "members": normalized_members,
    }
    if "require_open_shift_count_less_than_member_count" in config:
        normalized["config"]["require_open_shift_count_less_than_member_count"] = bool(
            config["require_open_shift_count_less_than_member_count"]
        )
    return normalized


def get_open_shift_list(data: dict) -> List[Tuple[str, str]]:
    shifts = []
    for day in DAYS:
        for slot in SLOTS:
            if data["config"]["open_shifts"][day][slot] == 1:
                shifts.append((day, slot))
    return shifts


def member_can_take_shift(member: dict, day: str, slot: str) -> bool:
    if day in member["ban_days"]:
        return False
    return member["schedule"][day][slot] == 0


def build_candidate_map(data: dict, shifts: List[Tuple[str, str]]) -> Dict[Tuple[str, str], List[str]]:
    candidate_map: Dict[Tuple[str, str], List[str]] = {}
    for day, slot in shifts:
        candidate_map[(day, slot)] = [
            member["name"] for member in data["members"] if member_can_take_shift(member, day, slot)
        ]
    return candidate_map


def run_matching(data: dict, shifts: List[Tuple[str, str]], candidate_map: Dict[Tuple[str, str], List[str]], cap: int) -> dict:
    members = data["members"]
    source = 0
    shift_offset = 1
    member_offset = shift_offset + len(shifts)
    sink = member_offset + len(members)
    mcmf = MinCostMaxFlow(sink + 1)

    shift_index = {shift: idx for idx, shift in enumerate(shifts)}
    member_index = {member["name"]: idx for idx, member in enumerate(members)}

    for shift, idx in shift_index.items():
        mcmf.add_edge(source, shift_offset + idx, 1, 0)
    for shift, idx in shift_index.items():
        for member_name in candidate_map[shift]:
            mcmf.add_edge(shift_offset + idx, member_offset + member_index[member_name], 1, 0)
    for member in members:
        member_node = member_offset + member_index[member["name"]]
        base_cost = PRIORITY_COST[member["priority"]]
        for repeat_layer in range(cap):
            mcmf.add_edge(member_node, sink, 1, repeat_layer * REPEAT_LAYER_COST + base_cost)

    flow, total_cost = mcmf.min_cost_max_flow(source, sink)

    assignments = []
    matched_counts = {member["name"]: 0 for member in members}
    for shift, idx in shift_index.items():
        node = shift_offset + idx
        matched_person = None
        for edge in mcmf.graph[node]:
            if member_offset <= edge.to < sink and edge.cap == 0:
                matched_person = members[edge.to - member_offset]["name"]
                break
        if matched_person is not None:
            matched_counts[matched_person] += 1
            assignments.append({"day": shift[0], "slot": shift[1], "person": matched_person})

    assignments.sort(key=lambda x: (x["day"], x["slot"], x["person"]))
    return {"flow": flow, "total_cost": total_cost, "assignments": assignments, "member_assignment_counts": matched_counts}


def analyze_capacity(data: dict, shifts: List[Tuple[str, str]], candidate_map: Dict[Tuple[str, str], List[str]], requested_cap: int) -> dict:
    total_open_shifts = len(shifts)
    if total_open_shifts == 0:
        return {
            "requested": requested_cap,
            "effective": 1,
            "recommended": 1,
            "advice_type": "ok",
            "message": "当前没有开启班次，无需调整每人最多排班次数。",
            "flow_at_requested": 0,
            "max_possible_if_more": 0,
            "full_possible_if_more": True,
            "flow_by_cap": {"1": 0},
        }

    max_cap_to_try = total_open_shifts
    capped_request = min(requested_cap, max_cap_to_try)
    flow_by_cap: Dict[int, int] = {}
    cost_by_cap: Dict[int, int] = {}
    for cap in range(1, max_cap_to_try + 1):
        result = run_matching(data, shifts, candidate_map, cap)
        flow_by_cap[cap] = result["flow"]
        cost_by_cap[cap] = result["total_cost"]

    flow_at_requested = flow_by_cap[capped_request]
    max_possible_if_more = flow_by_cap[max_cap_to_try]
    min_cap_for_requested_flow = min(cap for cap, flow in flow_by_cap.items() if flow >= flow_at_requested)
    min_cap_for_best_possible = min(cap for cap, flow in flow_by_cap.items() if flow >= max_possible_if_more)
    full_caps = [cap for cap, flow in flow_by_cap.items() if flow == total_open_shifts]
    min_cap_for_full = min(full_caps) if full_caps else None
    full_possible_if_more = min_cap_for_full is not None

    if flow_at_requested < max_possible_if_more:
        recommended = min_cap_for_full if min_cap_for_full is not None else min_cap_for_best_possible
        target_text = "排满全部班次" if min_cap_for_full is not None else "达到当前约束下的最大可行排班"
        advice_type = "too_small"
        message = (
            f"当前每人最多 {requested_cap} 次偏小，只能排 {flow_at_requested}/{total_open_shifts} 个开班时段；"
            f"建议至少设为 {recommended} 次，才能{target_text}。"
        )
        effective = min_cap_for_requested_flow
    elif requested_cap > min_cap_for_requested_flow:
        recommended = min_cap_for_requested_flow
        advice_type = "too_large"
        message = (
            f"当前每人最多 {requested_cap} 次偏大；设为 {recommended} 次已经能得到同样的排班填充效果"
            f"（{flow_at_requested}/{total_open_shifts}）。"
        )
        effective = recommended
    elif max_possible_if_more < total_open_shifts:
        recommended = min_cap_for_best_possible
        advice_type = "impossible_even_with_more"
        message = (
            f"即使继续提高每人最多排班次数，当前课表/禁排约束下最多也只能排 "
            f"{max_possible_if_more}/{total_open_shifts} 个开班时段；剩余时段只能空班。"
        )
        effective = min_cap_for_requested_flow
    else:
        recommended = requested_cap
        advice_type = "ok"
        message = f"当前每人最多 {requested_cap} 次设置合理。"
        effective = min_cap_for_requested_flow

    return {
        "requested": requested_cap,
        "effective": effective,
        "recommended": recommended,
        "advice_type": advice_type,
        "message": message,
        "flow_at_requested": flow_at_requested,
        "max_possible_if_more": max_possible_if_more,
        "full_possible_if_more": full_possible_if_more,
        "min_cap_for_full": min_cap_for_full,
        "min_cap_for_best_possible": min_cap_for_best_possible,
        "flow_by_cap": {str(k): v for k, v in flow_by_cap.items()},
    }


def solve_schedule(data: dict) -> dict:
    data = normalize_input(data)
    members = data["members"]
    shifts = get_open_shift_list(data)
    total_members = len(members)
    total_open_shifts = len(shifts)
    requested_cap = data["config"]["max_shifts_per_member"]
    candidate_map = build_candidate_map(data, shifts)
    capacity_advice = analyze_capacity(data, shifts, candidate_map, requested_cap)
    matching = run_matching(data, shifts, candidate_map, capacity_advice["effective"])
    total_cost = matching["total_cost"]
    assignments = matching["assignments"]
    member_assignment_counts = matching["member_assignment_counts"]

    duty_table = {day: {slot: None for slot in SLOTS} for day in DAYS}
    closed_shifts = []
    for day in DAYS:
        for slot in SLOTS:
            if data["config"]["open_shifts"][day][slot] == 0:
                duty_table[day][slot] = "不开班"
                closed_shifts.append({"day": day, "slot": slot})

    for item in assignments:
        duty_table[item["day"]][item["slot"]] = item["person"]

    unfilled_shifts = []
    for day, slot in shifts:
        if duty_table[day][slot] is None:
            unfilled_shifts.append({"day": day, "slot": slot})
            duty_table[day][slot] = "空班"

    no_candidate_shifts = [
        {"day": day, "slot": slot}
        for (day, slot), candidates in candidate_map.items()
        if len(candidates) == 0
    ]

    unassigned_members = [name for name, count in member_assignment_counts.items() if count == 0]
    assigned_members = {name for name, count in member_assignment_counts.items() if count > 0}
    all_filled = len(unfilled_shifts) == 0
    fill_rate = round((len(assignments) / total_open_shifts), 4) if total_open_shifts else 1.0
    max_shifts_used = max(member_assignment_counts.values()) if member_assignment_counts else 0

    warnings = []
    if "require_open_shift_count_less_than_member_count" in data["config"]:
        warnings.append(
            "字段 require_open_shift_count_less_than_member_count 已废弃。程序现在总是返回最大可行排班；即使无法全排满，也不会因此报错。"
        )
    if capacity_advice["advice_type"] != "ok":
        warnings.append(capacity_advice["message"])

    if all_filled:
        base_message = "已求得最大可行排班（全部填满）"
    elif len(assignments) == 0:
        base_message = "已求得最大可行排班（没有任何可分配班次）"
    else:
        base_message = "已求得最大可行排班（仅部分填满）"

    return {
        "success": True,
        "all_filled": all_filled,
        "status": "full" if all_filled else ("empty" if len(assignments) == 0 else "partial"),
        "message": base_message,
        "capacity_advice": capacity_advice,
        "summary": {
            "total_members": total_members,
            "open_shift_count": total_open_shifts,
            "assigned_shift_count": len(assignments),
            "unfilled_shift_count": len(unfilled_shifts),
            "closed_shift_count": len(closed_shifts),
            "unassigned_member_count": len(unassigned_members),
            "fill_rate": fill_rate,
            "max_shifts_per_member_requested": requested_cap,
            "max_shifts_per_member_effective": capacity_advice["effective"],
            "recommended_max_shifts_per_member": capacity_advice["recommended"],
            "max_shifts_used_in_result": max_shifts_used,
            "max_possible_if_more": capacity_advice["max_possible_if_more"],
            "full_possible_if_more": capacity_advice["full_possible_if_more"],
            "max_possible_by_headcount": min(total_open_shifts, total_members * requested_cap),
            "optimization_cost": total_cost,
            "priority_summary": {
                "prefer_more_assigned_shifts": sum(
                    member_assignment_counts[member["name"]]
                    for member in members
                    if member["priority"] == "prefer_more"
                ),
                "prefer_less_assigned_shifts": sum(
                    member_assignment_counts[member["name"]]
                    for member in members
                    if member["priority"] == "prefer_less"
                ),
                "prefer_more_assigned_members": sum(
                    1
                    for member in members
                    if member["priority"] == "prefer_more" and member["name"] in assigned_members
                ),
                "prefer_less_assigned_members": sum(
                    1
                    for member in members
                    if member["priority"] == "prefer_less" and member["name"] in assigned_members
                ),
            },
        },
        "assignments": assignments,
        "member_assignment_counts": member_assignment_counts,
        "duty_table": duty_table,
        "closed_shifts": closed_shifts,
        "no_candidate_shifts": no_candidate_shifts,
        "unfilled_shifts": unfilled_shifts,
        "unassigned_members": unassigned_members,
        "warnings": warnings,
        "normalized_input": data,
    }


def export_excel(result: dict, output_path: str) -> None:
    if Workbook is None:
        raise RuntimeError("当前环境未安装 openpyxl，无法导出 Excel")

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"

    title_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="EEF5FB")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    ws_summary["A1"] = "排班结果汇总"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = title_fill
    summary_rows = [
        ("程序运行成功", "是" if result["success"] else "否"),
        ("是否全部排满", "是" if result["all_filled"] else "否"),
        ("状态", result["status"]),
        ("说明", result["message"]),
        ("排班次数建议", result["capacity_advice"]["message"]),
        ("成员总数", result["summary"]["total_members"]),
        ("开班总数", result["summary"]["open_shift_count"]),
        ("已排班次数", result["summary"]["assigned_shift_count"]),
        ("未填满班次数", result["summary"]["unfilled_shift_count"]),
        ("填充率", result["summary"]["fill_rate"]),
        ("用户设置每人最多", result["summary"]["max_shifts_per_member_requested"]),
        ("实际使用上限", result["summary"]["max_shifts_per_member_effective"]),
        ("建议每人最多", result["summary"]["recommended_max_shifts_per_member"]),
        ("结果中单人最多实际排", result["summary"]["max_shifts_used_in_result"]),
    ]
    for i, (k, v) in enumerate(summary_rows, start=3):
        ws_summary[f"A{i}"] = k
        ws_summary[f"B{i}"] = v
        ws_summary[f"A{i}"].font = header_font

    ws_table = wb.create_sheet("DutyTable")
    ws_table["A1"] = "值班总表"
    ws_table["A1"].font = title_font
    ws_table["A1"].fill = title_fill
    headers = ["日期/班次"] + SLOTS
    for col, text in enumerate(headers, start=1):
        cell = ws_table.cell(row=3, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for r, day in enumerate(DAYS, start=4):
        ws_table.cell(row=r, column=1, value=DAY_NAME[day]).fill = header_fill
        ws_table.cell(row=r, column=1).font = header_font
        ws_table.cell(row=r, column=1).alignment = center
        for c, slot in enumerate(SLOTS, start=2):
            cell = ws_table.cell(row=r, column=c, value=result["duty_table"][day][slot])
            cell.alignment = center

    ws_assign = wb.create_sheet("Assignments")
    ws_assign["A1"] = "已分配班次"
    ws_assign["A1"].font = title_font
    ws_assign["A1"].fill = title_fill
    for c, text in enumerate(["星期", "班次", "成员"], start=1):
        cell = ws_assign.cell(row=3, column=c, value=text)
        cell.fill = header_fill
        cell.font = header_font
    for r, item in enumerate(result["assignments"], start=4):
        ws_assign.cell(row=r, column=1, value=DAY_NAME[item["day"]])
        ws_assign.cell(row=r, column=2, value=item["slot"])
        ws_assign.cell(row=r, column=3, value=item["person"])

    ws_counts = wb.create_sheet("MemberCounts")
    ws_counts["A1"] = "成员排班次数"
    ws_counts["A1"].font = title_font
    ws_counts["A1"].fill = title_fill
    for c, text in enumerate(["成员", "排班次数"], start=1):
        cell = ws_counts.cell(row=3, column=c, value=text)
        cell.fill = header_fill
        cell.font = header_font
    for r, (name, count) in enumerate(result["member_assignment_counts"].items(), start=4):
        ws_counts.cell(row=r, column=1, value=name)
        ws_counts.cell(row=r, column=2, value=count)

    ws_unfilled = wb.create_sheet("UnfilledShifts")
    ws_unfilled["A1"] = "未填满班次"
    ws_unfilled["A1"].font = title_font
    ws_unfilled["A1"].fill = title_fill
    for c, text in enumerate(["星期", "班次"], start=1):
        cell = ws_unfilled.cell(row=3, column=c, value=text)
        cell.fill = header_fill
        cell.font = header_font
    for r, item in enumerate(result["unfilled_shifts"], start=4):
        ws_unfilled.cell(row=r, column=1, value=DAY_NAME[item["day"]])
        ws_unfilled.cell(row=r, column=2, value=item["slot"])

    for ws in wb.worksheets:
        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)) + 2)
        for col_idx, width in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 10), 42)

    wb.save(output_path)


BASE_DIR = Path(__file__).resolve().parent
EXPORT_DIR = BASE_DIR / "exports"
app = Flask(__name__) if Flask is not None else None

if app is not None:
    @app.route("/")
    def index():
        return render_template("index.html", days=DAYS, slots=SLOTS, day_name=DAY_NAME)

    @app.route("/api/schedule", methods=["POST"])
    def api_schedule():
        try:
            data = request.get_json(force=True)
            result = solve_schedule(data)
            EXPORT_DIR.mkdir(exist_ok=True)
            token = datetime.now().strftime("%Y%m%d_%H%M%S") + "_" + uuid.uuid4().hex[:8]
            json_name = f"schedule_{token}.json"
            json_path = EXPORT_DIR / json_name
            json_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
            downloads = {"json": f"/downloads/{json_name}"}
            if Workbook is not None:
                xlsx_name = f"schedule_{token}.xlsx"
                export_excel(result, str(EXPORT_DIR / xlsx_name))
                downloads["xlsx"] = f"/downloads/{xlsx_name}"
            return jsonify({"ok": True, "result": result, "downloads": downloads})
        except ValidationError as e:
            return jsonify({"ok": False, "error": str(e)}), 400
        except Exception as e:
            return jsonify({"ok": False, "error": f"服务器错误：{e}"}), 500

    @app.route("/downloads/<path:filename>")
    def download_file(filename: str):
        return send_from_directory(str(EXPORT_DIR), filename, as_attachment=True)


def run_cli(input_path: str, output_json: Optional[str], output_xlsx: Optional[str]) -> None:
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    result = solve_schedule(data)
    if output_json:
        Path(output_json).write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    if output_xlsx:
        export_excel(result, output_xlsx)
    print(json.dumps({"status": result["status"], "summary": result["summary"], "capacity_advice": result["capacity_advice"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="社团值班排班器 V4")
    parser.add_argument("input", nargs="?", help="输入 JSON 文件路径")
    parser.add_argument("--json", dest="output_json", help="输出 JSON 文件路径")
    parser.add_argument("--xlsx", dest="output_xlsx", help="输出 Excel 文件路径")
    parser.add_argument("--web", action="store_true", help="启动网页表单")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=5000)
    args = parser.parse_args()

    if args.web:
        if app is None:
            raise SystemExit("当前环境未安装 Flask，请先执行: pip install flask")
        app.run(host=args.host, port=args.port, debug=True)
    else:
        if not args.input:
            raise SystemExit("CLI 模式下必须提供 input.json，或者使用 --web 启动网页表单")
        run_cli(args.input, args.output_json, args.output_xlsx)
